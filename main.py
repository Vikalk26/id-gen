# This script provides a complete suite for designing and generating ID cards.
# It includes a GUI for visual layout configuration and a card generation function.

import json
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog, colorchooser
from PIL import Image, ImageTk, ImageDraw, ImageFont
import os
import openpyxl
from openpyxl.styles import Font

# --- Configuration Constants ---
CARD_WIDTH_MM = 85.60
CARD_HEIGHT_MM = 53.98
DPI = 300
PIXELS_PER_MM = DPI / 25.4
CARD_WIDTH_PX = int(CARD_WIDTH_MM * PIXELS_PER_MM)
CARD_HEIGHT_PX = int(CARD_HEIGHT_MM * PIXELS_PER_MM)

CONFIG_FILE = 'config.json'

class CardProductionApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Card Production Suite")
        self.geometry("1000x800")

        self.config_data = {}
        self.fields = {} # Dictionary to hold label widgets
        self.drag_data = {"item": None, "x": 0, "y": 0}
        self.selected_field_name = None
        
        self.field_type_var = tk.StringVar(value="text") # Default to text

        # Initialize widgets first, then load the config
        self.create_widgets()
        self.load_config(CONFIG_FILE)

    def load_config(self, filepath):
        """Loads configuration from a JSON file."""
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r') as f:
                    self.config_data = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError) as e:
                messagebox.showerror("Error", f"Failed to load config file: {e}")
                self.config_data = {}
        
        if 'background_image' not in self.config_data:
            self.config_data['background_image'] = ''
        
        # Ensure all fields are properly structured
        for key in list(self.config_data.keys()):
            if key in ['background_image', 'border_color', 'border_width']:
                continue
            if 'x_mm' not in self.config_data[key] or 'y_mm' not in self.config_data[key]:
                del self.config_data[key] # Remove malformed fields
        
        self.update_canvas()
        self.update_fields_listbox()

    def load_config_from_file(self):
        """Opens a file dialog to select and load a new configuration file."""
        filepath = filedialog.askopenfilename(
            title="Select Configuration File",
            filetypes=[("JSON Files", "*.json")]
        )
        if filepath:
            self.load_config(filepath)
            messagebox.showinfo("Success", f"Configuration loaded from {os.path.basename(filepath)}.")

    def save_config(self):
        """Saves current configuration to JSON file."""
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config_data, f, indent=2)
            messagebox.showinfo("Success", "Configuration saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config file: {e}")

    def create_widgets(self):
        """Initializes and lays out the GUI widgets."""
        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left panel for controls
        control_frame = tk.Frame(main_frame, width=250)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        tk.Label(control_frame, text="Current Fields", font=("Arial", 12, "bold")).pack(pady=5)
        self.fields_listbox = tk.Listbox(control_frame, height=10)
        self.fields_listbox.pack(fill=tk.X, expand=True)
        self.fields_listbox.bind("<<ListboxSelect>>", self.on_field_select)

        add_field_frame = tk.Frame(control_frame)
        add_field_frame.pack(pady=10)
        
        tk.Label(add_field_frame, text="New Field Name:").grid(row=0, column=0, sticky="W")
        self.new_field_entry = tk.Entry(add_field_frame)
        self.new_field_entry.grid(row=0, column=1, padx=5, sticky="E")
        
        tk.Label(add_field_frame, text="Field Type:").grid(row=1, column=0, sticky="W")
        
        type_frame = tk.Frame(add_field_frame)
        type_frame.grid(row=1, column=1, sticky="W")
        tk.Radiobutton(type_frame, text="Text", variable=self.field_type_var, value="text").pack(side=tk.LEFT)
        tk.Radiobutton(type_frame, text="Image", variable=self.field_type_var, value="image").pack(side=tk.LEFT)
        
        tk.Button(add_field_frame, text="Add", command=self.add_field).grid(row=2, columnspan=2, pady=5)
        
        tk.Button(control_frame, text="Delete Selected Field", command=self.delete_field).pack(fill=tk.X, pady=5)
        
        # File management buttons
        tk.Button(control_frame, text="Save Configuration", command=self.save_config).pack(fill=tk.X, pady=10)
        tk.Button(control_frame, text="Load Configuration", command=self.load_config_from_file).pack(fill=tk.X, pady=5)
        tk.Button(control_frame, text="Generate Sample .xlsx", command=self.generate_sample_xlsx).pack(fill=tk.X, pady=5)
        
        # New "Generate Cards" button
        tk.Button(control_frame, text="Generate Cards", command=self.generate_cards).pack(fill=tk.X, pady=15, side=tk.BOTTOM)
        
        tk.Label(control_frame, text="Background Image:").pack(pady=(20, 0))
        self.bg_path_label = tk.Label(control_frame, text=self.config_data.get('background_image', 'None'), wraplength=200)
        self.bg_path_label.pack()
        tk.Button(control_frame, text="Select Image", command=self.select_background_image).pack(fill=tk.X, pady=5)

        # Right panel for the canvas
        canvas_frame = tk.Frame(main_frame, relief=tk.RAISED, borderwidth=1)
        canvas_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # We now use a fixed-size canvas to ensure accurate representation
        self.canvas = tk.Canvas(canvas_frame, width=CARD_WIDTH_PX, height=CARD_HEIGHT_PX, bg="white")
        self.canvas.pack(padx=20, pady=20)
        
        self.canvas.bind("<Button-1>", self.on_press)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        
        self.canvas.bind("<Button-3>", self.on_field_right_click)

        # Status bar
        self.status_label = tk.Label(self, text="Ready.", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)
        
    def update_canvas(self):
        """Clears and redraws the canvas with the current image and fields."""
        self.canvas.delete("all")
        
        try:
            bg_path = self.config_data.get('background_image')
            if bg_path and os.path.exists(bg_path):
                img = Image.open(bg_path)
                img = img.resize((CARD_WIDTH_PX, CARD_HEIGHT_PX), Image.Resampling.LANCZOS)
            else:
                img = Image.new("RGB", (CARD_WIDTH_PX, CARD_HEIGHT_PX), "white")
            
            self.tk_image = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, image=self.tk_image, anchor=tk.NW)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load background image: {e}")
            self.tk_image = None
            
        # Draw a border
        self.canvas.create_rectangle(0, 0, CARD_WIDTH_PX, CARD_HEIGHT_PX, outline="black", width=2)
            
        for field_name, props in self.config_data.items():
            if field_name in ['background_image', 'border_color', 'border_width']:
                continue
            
            try:
                # Use a simple conversion now that the canvas size is fixed
                x_mm = props['x_mm']
                y_mm = props['y_mm']
                x_px = int(x_mm * PIXELS_PER_MM)
                y_px = int(y_mm * PIXELS_PER_MM)
                
                field_type = props.get('type', 'text')
                
                if field_type == 'text':
                    # Check for default font and color
                    font_size = props.get('font_size', 12)
                    color = props.get('color', [0, 0, 0])
                    max_width_mm = props.get('max_width_mm', 50)
                    max_width_px = int(max_width_mm * PIXELS_PER_MM)
                    
                    fill_color = f"#{color[0]:02x}{color[1]:02x}{color[2]:02x}"
                    
                    self.canvas.create_text(x_px, y_px, text=field_name, fill=fill_color, font=("Arial", font_size), anchor=tk.NW, tags=(field_name, 'text_field'), width=max_width_px)
                
                elif field_type == 'image':
                    width_mm = props.get('width_mm', 25)
                    height_mm = props.get('height_mm', 30)
                    width_px = int(width_mm * PIXELS_PER_MM)
                    height_px = int(height_mm * PIXELS_PER_MM)
                    
                    self.canvas.create_rectangle(x_px, y_px, x_px + width_px, y_px + height_px, outline="black", width=2, tags=(field_name, 'image_field'))
                    self.canvas.create_text(x_px + width_px / 2, y_px + height_px / 2, text=field_name, fill="black", font=("Arial", 12, "italic"), tags=(field_name, 'image_text'))
                
            except KeyError as e:
                print(f"Skipping malformed field '{field_name}': missing key {e}")

    def update_fields_listbox(self):
        """Updates the listbox with current field names."""
        self.fields_listbox.delete(0, tk.END)
        for field_name in self.config_data.keys():
            if field_name not in ['background_image', 'border_color', 'border_width']:
                self.fields_listbox.insert(tk.END, field_name)

    def on_field_select(self, event):
        """Highlights the selected field on the canvas."""
        selected_index = self.fields_listbox.curselection()
        if not selected_index:
            return
        
        selected_field = self.fields_listbox.get(selected_index[0])
        self.canvas.itemconfig(selected_field, fill="blue")
        
        # Reset color of other fields
        for field_name in self.config_data.keys():
            if field_name not in ['background_image', 'border_color', 'border_width', selected_field]:
                if self.config_data[field_name].get('type') == 'text':
                    color = self.config_data.get(field_name, {}).get('color', [0, 0, 0])
                    fill_color = f"#{color[0]:02x}{color[1]:02x}{color[2]:02x}"
                    self.canvas.itemconfig(field_name, fill=fill_color)

    def on_press(self, event):
        """Starts a drag operation."""
        closest_item = self.canvas.find_closest(event.x, event.y)[0]
        tags = self.canvas.gettags(closest_item)
        
        # Check if the closest item belongs to a field
        if tags and tags[0] in self.config_data:
            field_name = tags[0]
            field_type = self.config_data[field_name].get('type')
            
            # Find the main tag to drag (either text or image rectangle)
            if field_type == 'text':
                drag_tag = field_name
            elif field_type == 'image':
                drag_tag = field_name
                
            self.drag_data["item"] = drag_tag
            self.drag_data["x"] = event.x
            self.drag_data["y"] = event.y

    def on_drag(self, event):
        """Continues the drag operation."""
        if self.drag_data["item"]:
            field_name = self.drag_data["item"]
            dx = event.x - self.drag_data["x"]
            dy = event.y - self.drag_data["y"]
            self.canvas.move(field_name, dx, dy)
            self.drag_data["x"] = event.x
            self.drag_data["y"] = event.y
            self.update_status()

    def on_release(self, event):
        """Ends the drag operation and updates the config."""
        if self.drag_data["item"]:
            # Get new position in canvas coordinates
            x_px, y_px = self.canvas.coords(self.drag_data["item"])
            
            # Convert back to mm and save to config using the fixed scale
            x_mm = x_px / PIXELS_PER_MM
            y_mm = y_px / PIXELS_PER_MM
            
            self.config_data[self.drag_data["item"]]['x_mm'] = round(x_mm, 2)
            self.config_data[self.drag_data["item"]]['y_mm'] = round(y_mm, 2)
            
            self.drag_data["item"] = None
            self.update_status()

    def update_status(self):
        """Updates the status bar with current coordinates."""
        if self.drag_data["item"]:
            x_px, y_px = self.canvas.coords(self.drag_data["item"])
            x_mm = x_px / PIXELS_PER_MM
            y_mm = y_px / PIXELS_PER_MM
            self.status_label.config(text=f"Dragging: {self.drag_data['item']} - Pos: ({x_mm:.2f} mm, {y_mm:.2f} mm)")
        else:
            self.status_label.config(text="Ready.")

    def add_field(self):
        """Adds a new text or image field to the config and canvas."""
        field_name = self.new_field_entry.get().strip()
        field_type = self.field_type_var.get()
        
        if not field_name:
            messagebox.showerror("Error", "Field name cannot be empty.")
            return
        if field_name in self.config_data:
            messagebox.showerror("Error", f"Field '{field_name}' already exists.")
            return
        
        if field_type == "text":
            self.config_data[field_name] = {
                "type": "text",
                "x_mm": 5,
                "y_mm": 5,
                "font_path": "arial.ttf",
                "font_size": 14,
                "color": [0, 0, 0],
                "max_width_mm": 50
            }
        elif field_type == "image":
            self.config_data[field_name] = {
                "type": "image",
                "x_mm": 5,
                "y_mm": 5,
                "width_mm": 25,
                "height_mm": 30,
                "border_radius_px": 20
            }

        self.update_canvas()
        self.update_fields_listbox()
        self.new_field_entry.delete(0, tk.END)

    def delete_field(self):
        """Deletes the selected text field from the config."""
        selected_index = self.fields_listbox.curselection()
        if not selected_index:
            messagebox.showwarning("Warning", "No field selected.")
            return
        
        field_name = self.fields_listbox.get(selected_index[0])
        if messagebox.askyesno("Delete Field", f"Are you sure you want to delete '{field_name}'?"):
            del self.config_data[field_name]
            self.update_canvas()
            self.update_fields_listbox()

    def generate_sample_xlsx(self):
        """Generates a sample XLSX file based on current field names."""
        field_names = [key for key in self.config_data.keys() if key not in ['background_image', 'border_color', 'border_width']]
        
        if not field_names:
            messagebox.showwarning("Warning", "No fields to export. Please add fields first.")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Sample Excel File"
        )
        
        if not filepath:
            return # User cancelled
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write headers with bold font
            header_font = Font(bold=True)
            for col_num, header in enumerate(field_names, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
            
            # Add a few rows of sample data
            sample_data_count = 5
            for row_num in range(2, sample_data_count + 2):
                for col_num, field_name in enumerate(field_names, 1):
                    field_type = self.config_data[field_name].get('type', 'text')
                    if field_type == 'image':
                        ws.cell(row=row_num, column=col_num, value=os.path.abspath(f"sample_photo_{row_num-1}.png"))
                    else:
                        ws.cell(row=row_num, column=col_num, value=f"Sample {field_name.replace('_', ' ').title()} {row_num-1}")

            # Auto-size columns for readability
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            messagebox.showinfo("Success", f"Sample XLSX file saved to: {os.path.basename(filepath)}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")

    def on_field_right_click(self, event):
        """
        Displays a context menu for the right-clicked text field.
        """
        closest_item = self.canvas.find_closest(event.x, event.y)[0]
        tags = self.canvas.gettags(closest_item)
        if tags and tags[0] in self.config_data:
            self.selected_field_name = tags[0]
            self.create_field_menu().post(event.x_root, event.y_root)
        
    def create_field_menu(self):
        """Creates the right-click context menu for a field."""
        menu = tk.Menu(self, tearoff=0)
        field_props = self.config_data[self.selected_field_name]
        field_type = field_props.get('type', 'text')
        
        if field_type == 'text':
            menu.add_command(label="Change Font Size", command=self.change_font_size)
            menu.add_command(label="Change Color", command=self.change_color)
            menu.add_command(label="Change Max Width", command=self.change_text_width)
        elif field_type == 'image':
            menu.add_command(label="Change Image Size", command=self.change_image_size)
            menu.add_command(label="Change Border Radius", command=self.change_border_radius)

        return menu

    def change_font_size(self):
        """Prompts the user to change the font size of the selected field."""
        if not self.selected_field_name:
            return
        
        current_size = self.config_data[self.selected_field_name]['font_size']
        new_size = simpledialog.askinteger("Font Size", "Enter new font size:", initialvalue=current_size, minvalue=1)
        if new_size is not None:
            self.config_data[self.selected_field_name]['font_size'] = new_size
            self.update_canvas()

    def change_color(self):
        """Prompts the user to change the color of the selected field."""
        if not self.selected_field_name:
            return
        
        current_color = self.config_data[self.selected_field_name]['color']
        hex_color = f"#{current_color[0]:02x}{current_color[1]:02x}{current_color[2]:02x}"
        new_color = colorchooser.askcolor(initialcolor=hex_color)
        if new_color:
            rgb_tuple = new_color[0]
            self.config_data[self.selected_field_name]['color'] = [int(c) for c in rgb_tuple]
            self.update_canvas()

    def change_text_width(self):
        """Prompts the user to change the max width of the selected text field."""
        if not self.selected_field_name:
            return
        
        current_width = self.config_data[self.selected_field_name].get('max_width_mm', 50)
        new_width = simpledialog.askinteger("Max Width (mm)", "Enter new maximum width in mm:", initialvalue=current_width, minvalue=1)
        if new_width is not None:
            self.config_data[self.selected_field_name]['max_width_mm'] = new_width
            self.update_canvas()
            
    def change_image_size(self):
        """Prompts the user to change the dimensions of the selected image field."""
        if not self.selected_field_name:
            return
        
        current_width = self.config_data[self.selected_field_name].get('width_mm', 25)
        current_height = self.config_data[self.selected_field_name].get('height_mm', 30)
        
        new_width = simpledialog.askinteger("Image Width (mm)", "Enter new width in mm:", initialvalue=current_width, minvalue=1)
        if new_width is not None:
            self.config_data[self.selected_field_name]['width_mm'] = new_width
            
        new_height = simpledialog.askinteger("Image Height (mm)", "Enter new height in mm:", initialvalue=current_height, minvalue=1)
        if new_height is not None:
            self.config_data[self.selected_field_name]['height_mm'] = new_height

        self.update_canvas()
        
    def change_border_radius(self):
        """Prompts the user to change the border radius of the selected image field."""
        if not self.selected_field_name:
            return
        
        current_radius = self.config_data[self.selected_field_name].get('border_radius_px', 20)
        new_radius = simpledialog.askinteger("Border Radius (px)", "Enter new border radius in pixels:", initialvalue=current_radius, minvalue=0)
        if new_radius is not None:
            self.config_data[self.selected_field_name]['border_radius_px'] = new_radius
            self.update_canvas()

    def select_background_image(self):
        """Opens a file dialog to select a new background image."""
        filepath = filedialog.askopenfilename(
            filetypes=[("Image Files", "*.png *.jpg *.jpeg")]
        )
        if filepath:
            self.config_data['background_image'] = filepath
            self.bg_path_label.config(text=os.path.basename(filepath))
            self.update_canvas()
    
    def generate_cards(self):
        """Generates cards based on the config and a selected data source."""
        data_filepath = filedialog.askopenfilename(
            title="Select Data Source (Excel file)",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not data_filepath:
            return
        
        output_dir = filedialog.askdirectory(
            title="Select Output Directory to Save Cards"
        )
        if not output_dir:
            return

        try:
            wb = openpyxl.load_workbook(data_filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            valid_fields = {k: v for k, v in self.config_data.items() if k in headers and k not in ['background_image', 'border_color', 'border_width']}
            
            if not valid_fields:
                messagebox.showerror("Error", "No matching fields found in the Excel file and config.")
                return

            bg_img_path = self.config_data.get('background_image')
            base_img = None
            if bg_img_path and os.path.exists(bg_img_path):
                base_img = Image.open(bg_img_path).resize((CARD_WIDTH_PX, CARD_HEIGHT_PX), Image.Resampling.LANCZOS).convert("RGBA")
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                card_data = dict(zip(headers, row))
                
                if base_img:
                    card_img = base_img.copy()
                else:
                    card_img = Image.new("RGBA", (CARD_WIDTH_PX, CARD_HEIGHT_PX), "white")
                
                draw = ImageDraw.Draw(card_img)
                
                for field_name, props in valid_fields.items():
                    field_type = props.get('type', 'text')
                    x_px = int(props['x_mm'] * PIXELS_PER_MM)
                    y_px = int(props['y_mm'] * PIXELS_PER_MM)
                    
                    if field_type == 'text':
                        text_content = str(card_data.get(field_name, ''))
                        try:
                            font = ImageFont.truetype(props.get('font_path', 'arial.ttf'), props.get('font_size', 12))
                        except IOError:
                            font = ImageFont.load_default()
                            
                        color_rgb = tuple(props.get('color', [0, 0, 0]))
                        max_width_px = int(props.get('max_width_mm', 50) * PIXELS_PER_MM)
                        
                        wrapped_text = ""
                        current_line = ""
                        for word in text_content.split():
                            if draw.textlength(current_line + " " + word, font=font) <= max_width_px:
                                current_line += " " + word
                            else:
                                wrapped_text += current_line.strip() + "\n"
                                current_line = word
                        wrapped_text += current_line.strip()
                        
                        draw.text((x_px, y_px), wrapped_text, font=font, fill=color_rgb)
                    
                    elif field_type == 'image':
                        image_path = str(card_data.get(field_name, ''))
                        if os.path.exists(image_path):
                            image_width_px = int(props.get('width_mm', 25) * PIXELS_PER_MM)
                            image_height_px = int(props.get('height_mm', 30) * PIXELS_PER_MM)
                            border_radius = props.get('border_radius_px', 20)
                            
                            original_image = Image.open(image_path).resize((image_width_px, image_height_px), Image.Resampling.LANCZOS).convert("RGBA")
                            
                            # Create a mask for rounded corners
                            mask = Image.new('L', original_image.size, 0)
                            draw_mask = ImageDraw.Draw(mask)
                            draw_mask.rounded_rectangle((0, 0, image_width_px, image_height_px), fill=255, radius=border_radius)
                            original_image.putalpha(mask)
                            
                            # Create a border
                            border_width = 2
                            border_img = Image.new('RGBA', original_image.size, (255, 255, 255, 0))
                            draw_border = ImageDraw.Draw(border_img)
                            draw_border.rounded_rectangle((0, 0, image_width_px - 1, image_height_px - 1), outline="black", width=border_width, radius=border_radius)
                            
                            # Paste the image onto the card
                            card_img.paste(original_image, (x_px, y_px), original_image)
                            card_img.paste(border_img, (x_px, y_px), border_img)
                            
                        else:
                            print(f"Warning: Image file not found at {image_path}. Skipping.")
                            
                card_name = f"card_{row_num-1}.png"
                output_path = os.path.join(output_dir, card_name)
                card_img.save(output_path)
            
            messagebox.showinfo("Success", f"Successfully generated {ws.max_row - 1} cards in '{os.path.basename(output_dir)}'.")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

if __name__ == "__main__":
    try:
        from PIL import Image, ImageTk, ImageDraw, ImageFont
    except ImportError:
        print("Pillow not found. Please install it using: pip install Pillow")
        exit()

    try:
        import openpyxl
    except ImportError:
        print("openpyxl not found. Please install it using: pip install openpyxl")
        exit()
    
    try:
        import tkinter
    except ImportError:
        print("Tkinter not found. It is usually included with Python, but if not, you may need to install it with your system's package manager.")
        exit()
        
    app = CardProductionApp()
    app.mainloop()
